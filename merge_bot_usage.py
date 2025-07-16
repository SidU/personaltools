import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, Iterable

try:  # optional import used for Excel files
    from openpyxl.utils.exceptions import InvalidFileException
except Exception:  # pragma: no cover - openpyxl may not be installed
    InvalidFileException = Exception



def read_shortlist(path: Path) -> list[dict]:
    """Return rows from the bot shortlist CSV."""
    with path.open(newline="", encoding="utf-8-sig") as f:
        sample = f.read(1024)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        reader = csv.DictReader(f, dialect=dialect)
        rows = []
        for row in reader:
            row = {k: v for k, v in row.items() if k is not None}
            rows.append(row)
    return rows


def _read_usage_excel(path: Path, scope: str) -> Dict[str, float]:
    """Return usage data from an Excel file filtered by scope."""
    try:  # pragma: no cover - optional dependency
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - optional dependency
        raise SystemExit(
            "openpyxl is required to read Excel files. Install it via pip."
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = {h: i for i, h in enumerate(header)}

    # Accept both 'AppId' and 'AppID' as valid keys for app id
    app_id_key = None
    for key in ("AppId", "AppID"):
        if key in col_idx:
            app_id_key = key
            break
    required = {app_id_key, "Active Users", "App Scope"}
    if not required <= set(col_idx):
        missing = required - set(col_idx)
        raise ValueError(f"Missing columns in usage data: {', '.join(missing)}")

    usage: Dict[str, float] = {}
    for row in rows[1:]:
        if row[col_idx["App Scope"]] != scope:
            continue
        app_id = row[col_idx[app_id_key]]
        active_users = row[col_idx["Active Users"]]
        if app_id is None or active_users is None:
            continue
        try:
            val = float(active_users)
        except (TypeError, ValueError):
            continue
        if app_id not in usage or usage[app_id] < val:
            usage[app_id] = val
    return usage


def _read_usage_csv(path: Path, scope: str) -> Dict[str, float]:
    """Return usage data from a CSV file filtered by scope."""
    with path.open(newline="", encoding="utf-8-sig") as f:
        sample = f.read(1024)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        reader = csv.DictReader(f, dialect=dialect)

        # Accept both 'AppId' and 'AppID' as valid keys for app id
        fieldnames = set(reader.fieldnames or [])
        app_id_key = None
        for key in ("AppId", "AppID"):
            if key in fieldnames:
                app_id_key = key
                break
        required = {app_id_key, "Active Users", "App Scope"}
        if not required <= fieldnames:
            missing = required - fieldnames
            raise ValueError(
                f"Missing columns in usage data: {', '.join(missing)}"
            )

        usage: Dict[str, float] = {}
        for row in reader:
            row = {k: v for k, v in row.items() if k is not None}
            if row.get("App Scope") != scope:
                continue
            app_id = row.get(app_id_key)
            active_users = row.get("Active Users")
            if not app_id or not active_users:
                continue
            try:
                val = float(active_users)
            except (TypeError, ValueError):
                continue
            if app_id not in usage or usage[app_id] < val:
                usage[app_id] = val

            # If app-id == 5be2b320-a5b7-4221-893c-dee506e4e365 then print the usage
            if app_id == "5be2b320-a5b7-4221-893c-dee506e4e365":
                print(f"Usage for {app_id}: {val}")

    return usage


def read_usage(path: Path, scope: str) -> Dict[str, float]:
    """Return mapping of AppId to Active Users filtered by scope."""
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return _read_usage_csv(path, scope)

    # sniff first bytes to detect CSV disguised with another extension
    try:
        with path.open("rb") as f:
            magic = f.read(4)
    except FileNotFoundError:
        raise
    if magic[:2] != b"PK":  # XLSX files are zip archives starting with PK
        return _read_usage_csv(path, scope)

    try:
        return _read_usage_excel(path, scope)
    except InvalidFileException:
        # not a valid Excel file, fall back to CSV
        return _read_usage_csv(path, scope)


def merge_data(shortlist: Iterable[dict], usage: Dict[str, float]) -> list[dict]:
    """Return merged rows with Active Users column sorted by usage."""
    merged = []
    for row in shortlist:
        row = {k: v for k, v in row.items() if k is not None}
        app_id = row.get("App ID") or row.get("AppId")
        row["Active Users"] = usage.get(app_id, 0)
        merged.append(row)
    merged.sort(key=lambda r: r.get("Active Users", 0) or 0, reverse=True)
    return merged


def write_csv(rows: Iterable[dict], path: Path | None) -> None:
    """Write rows as CSV to the given path or stdout, excluding app descriptions."""
    if not rows:
        return
    # Exclude description fields
    exclude_fields = {"Description", "description", "longDescription", "shortDescription"}
    fieldnames = sorted({k for row in rows for k in row.keys() if k is not None and k not in exclude_fields})
    out_file = open(path, "w", newline="", encoding="utf-8") if path else None
    writer = csv.DictWriter(
        out_file or sys.stdout, fieldnames=fieldnames, extrasaction="ignore"
    )
    writer.writeheader()
    for row in rows:
        clean = {k: v for k, v in row.items() if k is not None and k not in exclude_fields}
        writer.writerow(clean)
    if out_file:
        out_file.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge bot shortlist with usage data")
    parser.add_argument("botshortlist", type=Path, help="Path to shortlist CSV")
    parser.add_argument(
        "usagedata",
        type=Path,
        help="Path to usage data (CSV or XLSX)"
    )
    parser.add_argument("--rankby", choices=["Teams", "Group", "Meetings", "Personal"], default="Teams", help="Scope to rank by")
    parser.add_argument("-o", "--output", type=Path, help="Output CSV path", default=Path("merged_output.csv"))
    args = parser.parse_args()

    shortlist = read_shortlist(args.botshortlist)
    usage = read_usage(args.usagedata, args.rankby)
    merged = merge_data(shortlist, usage)
    write_csv(merged, args.output)


if __name__ == "__main__":
    main()
